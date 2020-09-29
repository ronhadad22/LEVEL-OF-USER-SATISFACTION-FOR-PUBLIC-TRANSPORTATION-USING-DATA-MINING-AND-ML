import glob
import os
import json
import DB_DAL
import analays_tweets
import xlsxwriter
import csv
import re
import pandas as pd
# import prepare_model_fasttest
from openpyxl import load_workbook
import xlwings as xw

import Streamer
import convert_to_vec

EXCEL_NAME = "my_tweets.xlsx"
EXCEL_PATH = r"C:\Users\ronha\Downloads\israel_w\dup1 works with files"
SHEET_NAME = "Sheet1"
path = os.path.join(EXCEL_PATH, EXCEL_NAME)


def main():
    if not os.path.exists(path):
        workbook = xlsxwriter.Workbook(path)
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', 'id')
        worksheet.write('B1', 'text')
        worksheet.write('C1', 'geotag')
        worksheet.write('D1', 'time')
        workbook.close()
    print("dffffffffffffffff")
    json_data = get_tweet_from_file("excel")
    # path = os.path.join(EXCEL_PATH, EXCEL_NAME)
    # write_to_excel(path, "data1111111")
    convert_to_vec.main()
    return


def get_tweet_from_file(mode):
    os.chdir(r"C:\Users\ronha\Downloads\israel_w\dup1 works with files")# this excel file is the output, consists all the json files which carry the arrived twits ( relevant twits)
    # first = 0
    count = 0 # counts the relevant tweets
    countAll = 0# counts all tweets
    for file in glob.glob("*.json"):
        try:
            print(file)
            print()
            with open(file, "a") as fp:
                if not Streamer.check_last_char(file, ']'): #------------------------------------------תיקנתייי----------
                    fp.write("]")
            with open(file) as fp:
                json1_str = fp.read()
                json_data = json.loads(json1_str)  # list of tweets (every tweet is a dictionary)
                for tweet in json_data:

                    countAll += 1
                    if analays_tweets.basic_filter_by_truck(tweet) == 1: # check if the relevant words exists in tweet
                        print(tweet)
                        count += 1
                        if mode == "excel":
                            add_to_excel(tweet, count + 1)# add to excel only the tweets the with relevant words
                print("finish with this file ")
                print(count)
                print(countAll)

        except IOError as err:
            print('\033[93m' + "fail to open the file" + file + '\033[0m' + str(err))
            return (0)
    print(count)
    print(countAll)
    print("finish all")


"""def post_to_db(json_data):
        #upload the tweets to mongoDB
        for data_for_add in json_data:g
            DB_DAL.DB_Collection.addOne(data_for_add)
            print(data_for_add)
        DB_DAL.DB_Collection.addmany(json_data, 1)
        #analize tweets by truck and uploud it to the second_filter_DB
        #for tweet in json_data:
    #analays_tweets.filter.filter_file_by_truck(tweet, 1, 0)
    print("done")
"""


# def add_to_excl(tweet_for_enter):
#     row = 1
#     fp = "tweets_ready.xlsx"
#     wb = load_workbook(fp)
#     sheet = wb.active
#     sheet.cell(row=2, column=4).value = '1'
#     wb.save(fp)
#
#     with open('tweets_file3.csv', mode='w') as tweets_file:
#             fieldnames = ['id', 'text', 'geotag', 'relevant tweets']
#             print("1")
#             writer = csv.DictWriter(tweets_file, fieldnames=fieldnames)
#             _id = tweet_for_enter['id'] % 100000000000000
#             coordinates = str(tweet_for_enter['place']['bounding_box']['coordinates'])
#             writer.writeheader()
#             print("2")
#             writer.writerow({'id': _id, 'text': tweet_for_enter['text'], 'geotag': coordinates, 'relevant tweets': "TBD"})
#     tweets_file.close()
#     workbook = xlsxwriter.Workbook('tweets_ready.xlsx')
#     worksheet = workbook.add_worksheet()
#     worksheet.write('A1', 'id')
#     worksheet.write('B1', 'text')
#     worksheet.write('C1', 'geotag')
#     worksheet.write('D1', 'relevant tweets (yes - 1 , no - 2)')
#     print(tweet_for_enter)
#     coordinates = str(tweet_for_enter['place']['bounding_box']['coordinates'])
#     id = tweet_for_enter['id'] % 100000000000000
#     worksheet.write(row, 0, id)
#     worksheet.write(row, 1, tweet_for_enter['text'])
#     worksheet.write(row, 2, coordinates)
#     row += 1
#     workbook.close()
#     print("done")
def is_file_exists(path):
    return os.path.exists(path)


def create_excel(path):
    open(path, 'wb')


def read_excel_data(path):
    with open(path, 'r') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
        for row in spamreader:
            print("--------------------------------------------------")
            colums_in_row = row[0].split("\t")
            for col in colums_in_row:
                print(col)


def add_to_excel(tweet, _row):
    try:
        wb = load_workbook(path)
        sheet = wb.active
        # sheet.cell(row=_row, column=2).value = "carmellll"
        if str(tweet['geo']) != 'None':
            coordinates = str(tweet['geo']['coordinates'])
        else:
            coordinates = 'None'
        id = int(tweet['id']) % 100000000000000
        sheet.cell(row=_row, column=1).value = id
        sheet.cell(row=_row, column=2).value = tweet['text']
        sheet.cell(row=_row, column=3).value = coordinates
        sheet.cell(row=_row, column=4).value = tweet['created_at']
        #     worksheet.write(row, 0, id)
        #     worksheet.write(row, 1, tweet_for_enter['text'])
        #     worksheet.write(row, 2, coordinates)
        wb.save(path)
        print("added to excel \n added to excel \nadded to excel \nadded to excel \n")


    except:
        print("not added to excel \n not added to excel \n not added to excel \n not added to excel \n")
        print('\033[93m' + "problem open file" + '\033[0m')


def add_to_excel_shani(tweet):
    path = os.path.join(EXCEL_PATH, EXCEL_NAME)
    if not is_file_exists(path):
        create_excel(path)


def get_sheet_max_row(writer, sheet_name):
    return writer.book[sheet_name].max_row


def add_data_to_cell(row, column, value):
    from xlutils.copy import copy
    from xlrd import open_workbook
    book_origin = open_workbook(EXCEL_PATH)
    book = copy(book_origin)
    sheet1 = book.get_sheet(0)
    sheet1.write(row, column, value)
    book.save(EXCEL_PATH)


def print_cells(writer, sheet_name):
    """

    :param writer:
    :type writer: pandas.io.excel._OpenpyxlWriter
    :param sheet_name:
    :return:
    """
    sheet = writer.book.sheet_by_name(sheet_name)
    print(sheet.row(1))
    # for row in writer.book[sheet_name].rows:

    # for cell in row:
    #     print("Value: " + str(cell.value))
    #     print("Column: " + str(cell.column))
    #     print("Row: " + str(cell.row))


is_sheet_exists = lambda writer: writer.sheets != {}


def load_excel_data_if_exists_for_relevant_sheet(writer, sheet_name):
    if is_sheet_exists(writer):
        writer.book = load_workbook(EXCEL_PATH)
        if SHEET_NAME not in writer.book.sheetnames:
            writer.book.create_sheet(sheet_name)
    else:
        writer.book.create_sheet(sheet_name)


def write_to_excel(path, value):
    writer = pd.ExcelWriter(path, engine='openpyxl')
    load_excel_data_if_exists_for_relevant_sheet(writer, SHEET_NAME)
    # add_data_to_cell(1, 1, "carmel")
    # max_shhet_row = writer.book[SHEET_NAME].max_row
    # print_cells(writer, SHEET_NAME)

    sheet = writer.book.sheet_by_name(SHEET_NAME)
    # value_to_insert_into_excel = pd.DataFrame.insert(1,1,'ssssssss')
    # sheet_max_row = get_sheet_max_row(writer, sheet_name)
    # value_to_insert_into_excel = pd.DataFrame.insert(1,1,'ssssssss',)
    # value_to_insert_into_excel.to_excel(writer, sheet_name, startrow=sheet_max_row)
    # writer.save()


if __name__ == '__main__':
    main()
