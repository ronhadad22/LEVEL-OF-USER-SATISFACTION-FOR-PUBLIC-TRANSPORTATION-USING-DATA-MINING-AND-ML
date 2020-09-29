import os
import json
import time
import datetime
from openpyxl import load_workbook
import xlsxwriter
import re
import queue

EXCEL_PATH = r"C:\Users\ronha\Downloads\israel_w\dup1 works with files"
SHEET_NAME = "Sheet1"
EXCEL_NAME = r"C:\Users\ronha\Downloads\israel_w\dup1 works with files\my_tweet" \
             r"s hour:"+str(datetime.datetime.fromtimestamp(time.time()).strftime('%Y_%m_%d'))+str(datetime.datetime.now().hour) + ".xlsx"
#counters to show the distribution
GEOTAG = 0
GEOTEXT = 0
GEOPLACE = 0
GEOUSER = 0
GEOWORST = 0
COUNTER = 0
GEObuild = 0
GEObart =0
GEOmuni = 0
GEOBOX_SAN_FRANCISCO = [-122.604, 37.564, -122.253, 37.866]



def geo_file_every_hour():
    # open the relevants files
    with open('muni.json') as json_file:
        muni = json.load(json_file)
        json_size = len(muni["features"])

    #suburb file
    EXCEL_PATH = r"C:\Users\ronha\Downloads"
    EXCEL_NAME = "geo_r.xlsx"
    path_3 = os.path.join(EXCEL_PATH, EXCEL_NAME)
    wb_3 = load_workbook(path_3)
    sheet_geo = wb_3.active

    #sf file
    EXCEL_PATH = r"C:\Users\ronha\Downloads"
    EXCEL_NAME = "san_geo.xlsx"
    path_4 = os.path.join(EXCEL_PATH, EXCEL_NAME)
    wb_4 = load_workbook(path_4)
    sheet_geo_san = wb_4.active

    #building file
    EXCEL_PATH = r"C:\Users\ronha\Downloads"
    EXCEL_NAME = "building_geo.xlsx"
    path_5 = os.path.join(EXCEL_PATH, EXCEL_NAME)
    wb_5 = load_workbook(path_5)
    sheet_building_geo = wb_5.active

    #bart station file
    EXCEL_PATH = r"C:\Users\ronha\Downloads"
    EXCEL_NAME = "bart.xlsx"
    path_6 = os.path.join(EXCEL_PATH, EXCEL_NAME)
    wb_6 = load_workbook(path_6)
    sheet_bart = wb_6.active

    # open all the files that we need to run by
    for filename in os.listdir(r'C:\Users\ronha\Downloads\israel_w\‏‏dup1 works with files2\temp_t'):
        EXCEL_PATH = r"C:\Users\ronha\Downloads\israel_w\‏‏dup1 works with files2\temp_t"
        EXCEL_NAME = filename
        path_1 = os.path.join(EXCEL_PATH, EXCEL_NAME)
        wb_1 = load_workbook(path_1)
        sheet = wb_1.active


        EXCEL_PATH = r"C:\Users\ronha\Downloads\israel_w\‏‏dup1 works with files2\temp_geo_final"
        EXCEL_NAME = " hour geo " + filename
        path_2 = os.path.join(EXCEL_PATH, EXCEL_NAME)
        #open the geo file that keep the extacted tweets with coordinate only
        with xlsxwriter.Workbook(path_2) as workbook_new:
            worksheet_new = workbook_new.add_worksheet()
            worksheet_new.write('A1', 'id')
            worksheet_new.write('B1', 'text')
            worksheet_new.write('C1', 'geotag')
            worksheet_new.write('D1', 'time')
            worksheet_new.write('E1', 'user_loc')
            worksheet_new.write('F1', 'cordinates')
            worksheet_new.write('G1', 'place')
            worksheet_new.write('H1', 'kind')
            worksheet_new.write('H1', 'kind')
            worksheet_new.write('J1', 'emo')
            worksheet_new.fileclosed
        wb_2 = load_workbook(path_2)
        sheet_new = wb_2.active


        index_new = 1

        for _row in range(2, sheet.max_row):
            count = 0
            #geo tag
            geotag = sheet.cell(row=_row, column=3).value


            if geotag != 'None' and geotag != 'None ' and geotag != None and type(geotag) != int:

                geotag = json.loads(geotag)

                # GEOBOX_SAN_FRANCISCO = [-122.604, 37.564, -122.253, 37.866]
                if geotag[0] > GEOBOX_SAN_FRANCISCO[1] and geotag[1] > GEOBOX_SAN_FRANCISCO[0] and geotag[0] < GEOBOX_SAN_FRANCISCO[3] and geotag[1]  < GEOBOX_SAN_FRANCISCO[2]:
                    count = count + 1;
                    global GEOTAG
                    GEOTAG = GEOTAG+1
                    sheet_new.cell(row=index_new + 1, column=8).value = "geo"



            text = sheet.cell(row=_row, column=2).value
            #adding emojicon to file
            emoticonP = [":-)", ":)", ":D", ":o)", ":]", ":3", ":c)", ":>", "=]", "8)"]

            emoticonN = [':[' ,':-(', ':(', ':-c',':c', ':-<', ':<', ':-[', ':[', ':{']
            for emo in emoticonP:
                if text.lower().find(emo) >= 0:
                    sheet_new.cell(row=index_new + 1, column=10).value = "pos"
                    break
            for emo in emoticonN:
                if text.lower().find(emo) >=0 :
                    sheet_new.cell(row=index_new + 1, column=10).value = "neg"
                    break

            # -----------------bart
            if count == 0:
                for i in range(2, sheet_bart.max_row):

                    if text != None and type(text) == str and sheet_bart.cell(row=i, column=1).value != None:
                        if text.lower().find(sheet_bart.cell(row=i, column=1).value.lower()) >= 0:
                            sheet_new.cell(row=index_new + 1, column=3).value = sheet_geo.cell(row=i, column=2).value
                            sheet_new.cell(row=index_new + 1, column=8).value = "bart_geo"
                            count = count + 1
                            global GEObart
                            GEObart = GEObart + 1
                            break
            #===============muni

            if count == 0:
                last_str=""
                for i in muni["features"]:
                    str_ = i["properties"]["stop_name"].rstrip()
                    # print(str.rstrip() == "16th St & Church St")
                    # print(str_)

                    x = str_
                    two_=0
                    if(str_.lower().find("&")>=0):
                        x = str_.split("&")
                        # print(str_.split("&"))
                        # print(x[1].lstrip())
                        # print(x[0])
                        two_=1

                        x = x[0]

                    if last_str == str(i["properties"]["stop_name"]) :
                        continue
                    # for x in range(len(x)):
                    #     x[1].lstrip()
                    #     vale.append()
                    # print()
                    if text != None and type(text) == str and i["properties"]["stop_name"] != None and text.lower().find("nyc") >= 0:
                        if two_== 0:
                            if text.lower().find(x.lower()) >= 0:  # text.lower().find(x.lstrip().lower()):
                                two_ = 3
                        if two_== 1:
                            if text.lower().find(x.lower()) >= 0 and text.lower().find(x.lstrip().lower()):
                                two_ = 3
                        if two_==3:

                            sheet_new.cell(row=index_new + 1, column=3).value = str(i["geometry"]["coordinates"])
                            sheet_new.cell(row=index_new + 1, column=8).value = "muni_geo"
                            count = count + 1
                            global GEOmuni
                            GEOmuni = GEOmuni + 1
                            break
                    last_str == str(i["properties"]["stop_name"])

            #----------------geo building
            if count == 0:
                for i in range(2, sheet_building_geo.max_row):

                    if text != None and type(text) == str and sheet_building_geo.cell(row=i, column=1).value != None:
                        # print(sheet_building_geo.cell(row=i, column=1).value)
                        if text.lower().find(sheet_building_geo.cell(row=i, column=1).value.lower()) >= 0:
                            sheet_new.cell(row=index_new + 1, column=3).value = "["+sheet_geo.cell(row=i, column=2).value+"]"
                            sheet_new.cell(row=index_new + 1, column=8).value = "building_geo"
                            count = count + 1
                            global GEObuild
                            GEObuild = GEObuild + 1
                            break
            # --------------------TEXT

            sheet_new.cell(row=index_new + 1, column=9).value = "Null"
            if count == 0:
                for i in range(2, sheet_geo.max_row):

                    if text != None and type(text) == str:
                        if sheet_geo.cell(row=i, column=1).value.lower() == "soma":
                            sheet_geo.cell(row=i, column=1).value="soma "
                        if text.lower().find(sheet_geo.cell(row=i, column=1).value.lower()) >= 0:

                            sheet_new.cell(row=index_new+1, column=3).value = "["+sheet_geo.cell(row=i, column=2).value+"]"
                            sheet_new.cell(row=index_new + 1, column=8).value = "text"
                            count = count + 1
                            global GEOTEXT
                            GEOTEXT = GEOTEXT+1
                            break
            # place
            if count == 0:
                if sheet.cell(row=1, column=7).value == "place":
                    print("place")
                    print(sheet.cell(row=1, column=7).value)
                    place = sheet.cell(row=_row, column=7).value
                    if place != "ko" and place != None and count == 0:

                        place = place.replace("\'", "\"")
                        if len(place) > 1000:
                            print(place)

                        while True:
                            try:

                                place = json.loads(place)
                                break
                            except BaseException as e:

                                print('\033[93m' + "Error on data: %s" % str(e) + '\033[0m')
                                x = [int(s) for s in str(e).split() if s.isdigit()]

                                place = place[:x[1] - 1] + place[x[1]:]

                                pass

                        geo_1 = place["bounding_box"]["coordinates"][0][0][0]
                        geo_2 = place["bounding_box"]["coordinates"][0][0][1]
                        geo_3 = place["bounding_box"]["coordinates"][0][2][0]
                        geo_4 = place["bounding_box"]["coordinates"][0][1][1]

                        if geo_1 > GEOBOX_SAN_FRANCISCO[0] and geo_2 > GEOBOX_SAN_FRANCISCO[1] and geo_3 < \
                                GEOBOX_SAN_FRANCISCO[
                                    2] and geo_4 < GEOBOX_SAN_FRANCISCO[3]:
                            sheet_new.cell(row=index_new + 1, column=3).value = "[" + str(
                                float(geo_1 + geo_3) / 2) + "," + str(float(geo_2 + geo_4) / 2) + "]"
                            sheet_new.cell(row=index_new + 1, column=8).value = "place"
                            count = count + 1
                            global GEOPLACE
                            GEOPLACE = GEOPLACE + 1
            #user---------------
            if count == 0:
                if sheet.cell(row=1, column=5).value == "user_loc":
                    user_loc = sheet.cell(row=_row, column=5).value
                    # sheet_new.cell(row=index_new + 1, column=10).value = 0
                    if user_loc != None and count == 0:
                        if user_loc.lower().find("san francisco") >= 0 or user_loc.lower().find(" bay ") >= 0 or user_loc.lower().find(" sf") >= 0:
                            sheet_new.cell(row=index_new+1, column=3).value = "user"
                            sheet_new.cell(row=index_new + 1, column=8).value = "user"
                            global GEOUSER
                            GEOUSER = GEOUSER + 1
                            count = count+1

            #the worst case
            if count == 0:
                text = sheet.cell(row=_row, column=2).value
                for i in range(2, sheet_geo_san.max_row):

                    if text != None and type(text) == str:
                        # print(text)
                        # print(sheet_geo_san.cell(row=i, column=1).value)
                        # print(type(text))
                        if text.lower().find(sheet_geo_san.cell(row=i, column=1).value.lower()) >= 0:
                            sheet_new.cell(row=index_new+1, column=3).value = "ron_re"
                            sheet_new.cell(row=index_new + 1, column=8).value = "worst"
                            count = count + 1
                            global GEOWORST
                            GEOWORST = GEOWORST + 1
                            break

            if count > 1:
                global COUNTER
                COUNTER = COUNTER + 1
                print("COUNTER"+str(COUNTER))
            if count > 0:
                index_new = index_new + 1
                sheet_new.cell(row=index_new, column=1).value = sheet.cell(row=_row, column=1).value
                sheet_new.cell(row=index_new, column=2).value = sheet.cell(row=_row, column=2).value
                sheet_new.cell(row=index_new, column=4).value = sheet.cell(row=_row, column=4).value
                sheet_new.cell(row=index_new, column=5).value = sheet.cell(row=_row, column=5).value
                sheet_new.cell(row=index_new, column=6).value = sheet.cell(row=_row, column=6).value
                sheet_new.cell(row=index_new, column=7).value = sheet.cell(row=_row, column=7).value

            print(count)
            print("GEOWORST: "+str(GEOWORST))
            print("GEOTAG: "+str(GEOTAG))
            print("GEOPLACE: "+str(GEOPLACE))
            print("GEOWORST: "+str(GEOWORST))
            print("GEOTEXT: " + str(GEOTEXT))
            print("GEOUSER: " + str(GEOUSER))
            print("GEObuild" + str(GEObuild))
            print("GEObart" + str(GEObart))
            print("GEOmuni" + str(GEOmuni))
            print("COUNTER" + str(COUNTER))

        wb_1.save(path_1)
        wb_2.save(path_2)
    wb_3.save(path_3)
    wb_4.save(path_4)
    wb_5.save(path_5)
    wb_6.save(path_6)

if __name__ == '__main__':

    geo_file_every_hour()