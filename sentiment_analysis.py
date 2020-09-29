import re
import xlrd
import os
import pandas as pd
from openpyxl import load_workbook
from textblob import TextBlob
import time
import datetime
#import fastText
from prepare_model_fasttest import tweet_cleaning_for_sentiment_analysis
import main
import applying_ML_algorithms


EXCEL_NAME = "sub_v0.csv"
EXCEL_PATH = r"D:\semester\final project\BASE CODE"
# SHEET_NAME = "Sheet1"
path = os.path.join(EXCEL_PATH, EXCEL_NAME)
SIZE_OF_EXCEL = 790
fp = r'D:\semester\final project\BASE CODE\files'

def get_tweet_from_excel(index):
    # read from excel file the text cell in given index.
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    return sheet.cell_value(index, 1)
def result_cal(result):
    if result > 0.5:
        result = 2 * (result - 0.5)
    elif result < 0.2 and result > -0.2:
        result = 0
    else:
        result = -1 + (1 / 1.5) * (result + 1)
    return result
def set_result_to_excel_TextBlob(index, result):
    # read from excel file the text cell in given index.
    try:
        wb = load_workbook(fp)
        print(fp)
        sheet = wb.active
        if result > 0.5:
            result = 2*(result-0.5)
        elif result < 0.2 and result > -0.2:
            result = 0
        else:
            result = -1+(1/1.5)*(result+1)
        sheet.cell(row=index, column=5).value = result
        print("result")
        print(result)
        if result > 0:
            sheet.cell(row=index, column=6).value = 'positive'
        if result == 0:
            sheet.cell(row=index, column=6).value = 'neutral'
        if result < 0:
            sheet.cell(row=index, column=6).value = 'negative'
        wb.save(fp)

    except:
        print('\033[93m' + "problem open file1" + '\033[0m')

def set_result_to_excel_fastTest(index, result):
    # read from excel file the text cell in given index.
    try:
        wb = load_workbook(fp)
        sheet = wb.active
        # print(result)
        sheet.cell(row=index, column=7).value = result[1][0]
        if result[0][0] == '__label__NEGATIVE':
            sheet.cell(row=index, column=8).value = 'negative'
        if result[0][0] == '__label__NEUTRAL':
            sheet.cell(row=index, column=8).value = 'neutral'
        if result[0][0] == '__label__POSITIVE':
            sheet.cell(row=index, column=8).value = 'positive'

        wb.save(fp)

    except:
        print('\033[93m' + "problem open file2" + '\033[0m')

def clean_tweet(tweet):
    """
    Utility function to clean tweet text by removing links, special characters
    using simple regex statements.
    """
    return ' '.join(re.sub("(@[A-Za-z0-9]+)|([^0-9A-Za-z \t])| (\w+:\ / \ / \S+)", " ", tweet).split())

def get_tweet_sentiment_with_TextBlob(tweet):
    """
    Utility function to classify sentiment of passed tweet
    using textblob's sentiment method
    """
    # create TextBlob object of passed tweet text
    clean_tweets = clean_tweet(tweet)
    #print("check: ", type(clean_tweets), clean_tweets)

    analysis = TextBlob(clean_tweets)
# set sentiment
    result = analysis.sentiment.polarity
  #  print("5555555555555555555")
  #  print(result)
    list_of_words = ['delay', 'late', 'wait']
    words_re = re.compile("|".join(list_of_words))
    if words_re.search(clean_tweets):
        result = -1
    words_re = re.compile("|".join('minutes'))
    if words_re.search(clean_tweets) and result > -0.6:
        result = result-0.4

    #print(result)
    return result

def get_tweet_sentiment_with_fastTest(tweet):
    modelPath = "C:/Users/Carmel/PycharmProjects/final_pro_new/model-en.ftz"
    model = fastText.load_model(modelPath)
    result = model.predict(tweet, k=4)

    #print(model.predict(["I'm very pleased with the new downtown Berkeley BART plaza. Nice entrances; nice bus shelters. Beautiful Berkeley."], k=3))
    return result


def main1(filename):
    date = datetime.datetime.fromtimestamp(time.time()).strftime('%Y_%m_%d')
    now = datetime.datetime.now()
    final_name = "final_" + str(now.hour) + "_" + str(now.minute) + "_" + str(now.second) + "__" + str(date) + ".xlsx"
    final_path = os.path.join(main.final_result_path, final_name)
    try:
        if os.path.exists(final_path):
            open(final_path, 'w').close()
        read_file = pd.read_csv(filename)  # read file subv0
        read_file.to_excel(final_path, index=None, header=True)
        wb = load_workbook(final_path)
        sheet = wb.active
        sheet.cell(row=1, column=5).value = 'TextBlob - certainty'
        sheet.cell(row=1, column=6).value = 'TextBlob - decision'
        # sheet.cell(row=1, column=7).value = 'fastTest - certainty'
        # sheet.cell(row=1, column=8).value = 'fastTest - decision'
        wb.save(final_path)
    except:
        print('\033[93m' + "problem opengggg file" + '\033[0m')
    #for index in range(1, 30):
    wb = load_workbook(final_path)
    sheet = wb.active
    max_row = sheet.max_row
    for index in range(2, max_row+1):
        tweet = sheet.cell(row=index, column=4)
        geo = sheet.cell(row=index, column=5).value
        tweet_after_clean = tweet_cleaning_for_sentiment_analysis(tweet.value)
        print("SENTIMENT_ANALYSIS: TWEET_AFTER_CLEAN = ", tweet_after_clean)
        #result = get_tweet_sentiment_with_fastTest(tweet_after_clean)
        # print(tweet_after_clean)
        # print(result)
        #set_result_to_excel_fastTest(index+2, result)
        result = get_tweet_sentiment_with_TextBlob(tweet.value)
        result = result_cal(result)
        sheet.cell(row=index, column=7).value = tweet_after_clean
        sheet.cell(row=index, column=8).value = geo
        sheet.cell(row=index, column=5).value = result

        if result > 0:
            sheet.cell(row=index, column=6).value = 'positive'
        if result == 0:
            sheet.cell(row=index, column=6).value = 'neutral'
        if result < 0:
            sheet.cell(row=index, column=6).value = 'negative'
        #set_result_to_excel_TextBlob(index, result)
        if index % 100 == 0:
            print(index)
    print("SENTIMENT_ANALYSIS: Finish")
    wb.save(final_path)


if __name__ == '__main__':
    main1(applying_ML_algorithms.filename)
