import os
from tweepy import OAuthHandler
from tweepy import Stream
from tweepy.streaming import StreamListener

import main
import sys
import DB_DAL
import analays_tweets
import twitter_credentials
import json
import time
import datetime
from openpyxl import load_workbook
import xlsxwriter
import geo_extract
import re
import queue

EXCEL_PATH = r"C:\Users\ronha\Downloads\israel_w\dup1 works with files"
SHEET_NAME = "Sheet1"
EXCEL_NAME = r"C:\Users\ronha\Downloads\israel_w\dup1 works with files\my_tweet" \
             r"s hour:"+str(datetime.datetime.fromtimestamp(time.time()).strftime('%Y_%m_%d'))+str(datetime.datetime.now().hour) + ".xlsx"

def add_to_excel(tweet, _row, st):
    try:

        wb = load_workbook(st)
        sheet = wb.active
        if str(tweet['geo']) != 'None':
            coordinates = str(tweet['geo']['coordinates'])
            sheet.cell(row=_row, column=9).value = "ron the"
        else:
            coordinates = 'None'
        if str(tweet['media_url']) != 'None':
            sheet.cell(row=_row, column=10).value = str(tweet['media_url'])
        else:
            sheet.cell(row=_row, column=10).value = "None"
        user_loc = 'None'
        if str(tweet["user"]["geo_enabled"]) != 'None':
            user_loc = str(tweet["user"]["location"])

        cor = 'None'
        if str(tweet["coordinates"]) != 'None':
            cor = str(tweet["coordinates"]["coordinates"])

        w = 'not'
        place = 'ko'
        if str(tweet["place"]) != 'None':
            place = str(tweet["place"])
            w = 'right'
            print("place:  ")

            print(str(tweet["place"]["full_name"]))
            print(str(tweet["place"]["country_code"]))
            print(str(tweet["place"]["country"]))
            print(str(tweet["place"]["bounding_box"]))

        id = int(tweet['id']) % 100000000000000
        sheet.cell(row=_row, column=1).value = id
        sheet.cell(row=_row, column=2).value = tweet['text']
        sheet.cell(row=_row, column=3).value = coordinates
        sheet.cell(row=_row, column=4).value = tweet['created_at']
        sheet.cell(row=_row, column=5).value = user_loc
        sheet.cell(row=_row, column=6).value = cor
        #sheet.cell(row=_row, column=8).value = str(tweet['source'])
        sheet.cell(row=_row, column=7).value = place

        #         #     worksheet.write(row, 0, id)
        #     worksheet.write(row, 1, tweet_for_enter['text'])
        #     worksheet.write(row, 2, coordinates)

        wb.save(st)


    except:
        print("not added to excel \n not added to excel \n not added to excel \n not added to excel \n")
        print('\033[93m' + "problem open file" + '\033[0m')


def save_file_every_hour():
    while 1:
        # if main.SAVE_FILE_FLAG == 1:
        st = datetime.datetime.fromtimestamp(time.time()).strftime('%Y_%m_%d')



        EXCEL_PATH = r"C:\Users\ronha\Downloads\israel_w\‏‏dup1 works with files2\temp"
        EXCEL_NAME = str(st) + " hour " + str(datetime.datetime.now().hour) + "first.xlsx"
        path_ = os.path.join(EXCEL_PATH, EXCEL_NAME)
        json_file_name = " "
        if datetime.datetime.now().hour == 0:
            json_file_name = str(st)+str(23) + "temp_t.json"
        else:
            json_file_name = str(st)+str(datetime.datetime.now().hour - 1) + "temp_t.json"


        print(json_file_name)
        EXCEL_PATH_ = r"C:\Users\ronha\Downloads\israel_w\dup1 works with files"
        EXCEL_NAME_ = json_file_name
        json_file_name = os.path.join(EXCEL_PATH_, EXCEL_NAME_)
        if os.path.exists(json_file_name):
            print("dddd")
        if os.path.exists(path_):
            print("dddddd")
        if datetime.datetime.now().minute > 0 and not(os.path.exists(path_)) and os.path.exists(json_file_name):


            with open(json_file_name) as fp:

                json1_str = fp.read()
                try:

                    json_data = json.loads(json1_str)  # list of tweets (every tweet is a dictionary)
                    count = 0
                    print(json_data)
                    workbook = xlsxwriter.Workbook(path_)
                    worksheet = workbook.add_worksheet()
                    worksheet.write('A1', 'id')
                    worksheet.write('B1', 'text')
                    worksheet.write('C1', 'geotag')
                    worksheet.write('D1', 'time')
                    worksheet.write('E1', 'user_loc')
                    worksheet.write('F1', 'cordinates')
                    worksheet.write('G1', 'place')
                    workbook.close()

                    for tweet in json_data:

                        if str(tweet[
                                  'source']) == r'<a href="http://twitter.com/download/android" rel="nofollow">Twitter for Android</a>' or str(tweet["coordinates"] != 'None') or str(
                                tweet[
                                    'source']) == r'<a href="https://mobile.twitter.com" rel="nofollow">Twitter Web App</a>' or str(
                                tweet[
                                    'source']) == '<a href="http://twitter.com/download/iphone" rel="nofollow">Twitter for iPhone</a>':
               #             if tweet[]
                            count += 1
                            add_to_excel(tweet, count + 1, path_)  # add to excel only the tweets the with relevant words
                # main.SAVE_FILE_FLAG = 0
                    print("save the temp file")
                    geo_extract.geo_file_every_hour()

                except ValueError as e:
                    print('\033[93m' + "json Error on data: %s" % str(e) + '\033[0m')
                    with open(json_file_name, 'a') as fp:
                        if os.stat(json_file_name).st_size != 0:
                            fp.write("]")
                            fix = 1
                    pass
        else:
            print("main.SAVE_FILE_FLAG: save to file")
            # print(main.SAVE_FILE_FLAG)
            print("sleeppp")
            time.sleep(10)

if __name__ == '__main__':
    save_file_every_hour()